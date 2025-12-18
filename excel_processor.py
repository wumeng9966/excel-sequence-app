# excel_processor.py (改进版)
import os
import time
import traceback
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import io

def get_sequence_from_website_alternative(input_value, retry_count=3):
    """
    使用 requests 替代 Selenium 从网站获取序列
    添加重试机制和错误处理
    
    Args:
        input_value: 输入值，格式如"chr4B:425000640-425000640"
        retry_count: 重试次数
    
    Returns:
        str: 获取的序列，如果失败则返回空字符串
    """
    for attempt in range(retry_count):
        try:
            # 网站 URL
            url = "http://202.194.139.32/getfasta/getfasta.php"
            
            # 构建 POST 请求的参数
            data = {
                'database': 'Chinese_Spring1.0.genome',
                'ID': input_value
            }
            
            # 添加请求头，模拟浏览器
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Content-Type': 'application/x-www-form-urlencoded'
            }
            
            # 发送请求，增加超时时间
            response = requests.post(
                url, 
                data=data, 
                headers=headers, 
                timeout=30,
                verify=False  # 如果有SSL证书问题
            )
            
            # 检查响应
            if response.status_code == 200:
                # 使用 BeautifulSoup 解析 HTML
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 查找序列部分
                seq_div = soup.find('pre', {'id': 'seq'})
                
                if seq_div:
                    # 获取序列文本
                    seq_text = seq_div.text
                    
                    # 提取序列部分（去掉第一行的标题）
                    lines = seq_text.split('\n')
                    if len(lines) > 1:
                        # 合并所有序列行（去掉第一行的标题）
                        sequence = ''.join(lines[1:])
                        if sequence and len(sequence) > 0:
                            return sequence.strip()
                    return ""
                else:
                    print(f"尝试 {attempt+1}/{retry_count}: 未找到序列元素")
                    if attempt < retry_count - 1:
                        time.sleep(1)  # 等待后重试
            else:
                print(f"尝试 {attempt+1}/{retry_count}: 请求失败，状态码: {response.status_code}")
                if attempt < retry_count - 1:
                    time.sleep(1)
                    
        except requests.exceptions.Timeout:
            print(f"尝试 {attempt+1}/{retry_count}: 请求超时")
            if attempt < retry_count - 1:
                time.sleep(2)
        except requests.exceptions.RequestException as e:
            print(f"尝试 {attempt+1}/{retry_count}: 请求异常: {str(e)}")
            if attempt < retry_count - 1:
                time.sleep(1)
        except Exception as e:
            print(f"尝试 {attempt+1}/{retry_count}: 未知错误: {str(e)}")
            if attempt < retry_count - 1:
                time.sleep(1)
    
    return ""

def process_excel_with_sequences(uploaded_file_content, max_rows=None):
    """
    处理Excel文件，为每一行的K列和O列获取序列
    
    Args:
        uploaded_file_content: 上传的Excel文件内容（bytes）
        max_rows: 最大处理行数，None表示处理所有行
    
    Returns:
        tuple: (成功处理的行数, 处理后的文件内容)
    """
    try:
        # 将上传的文件内容保存到内存中的 BytesIO 对象
        excel_file = io.BytesIO(uploaded_file_content)
        
        # 读取Excel文件
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        processed_count = 0
        success_count = 0
        
        # 确定要处理的行数
        total_rows = ws.max_row
        if max_rows and max_rows < total_rows:
            total_rows = max_rows
        
        # 处理每一行（从第1行开始）
        for row_idx in range(1, total_rows + 1):
            # ========== 处理K列 ==========
            k_value = ws.cell(row=row_idx, column=11).value  # K列是第11列
            if k_value:
                try:
                    # 转换输入值为字符串
                    input_str = str(k_value).strip()
                    
                    # 检查输入格式是否正确
                    if ":" in input_str and "-" in input_str:
                        sequence = get_sequence_from_website_alternative(input_str)
                        
                        if sequence:
                            ws.cell(row=row_idx, column=12, value=sequence)  # L列是第12列
                            success_count += 1
                        else:
                            ws.cell(row=row_idx, column=12, value="获取失败")
                    else:
                        ws.cell(row=row_idx, column=12, value="格式错误")
                except Exception as e:
                    ws.cell(row=row_idx, column=12, value="处理出错")
            else:
                ws.cell(row=row_idx, column=12, value="空值")
            
            # 短暂等待，避免对网站造成过大压力
            time.sleep(0.3)
            
            # ========== 处理O列 ==========
            o_value = ws.cell(row=row_idx, column=15).value  # O列是第15列
            if o_value:
                try:
                    # 转换输入值为字符串
                    input_str = str(o_value).strip()
                    
                    # 检查输入格式是否正确
                    if ":" in input_str and "-" in input_str:
                        sequence = get_sequence_from_website_alternative(input_str)
                        
                        if sequence:
                            ws.cell(row=row_idx, column=16, value=sequence)  # P列是第16列
                            success_count += 1
                        else:
                            ws.cell(row=row_idx, column=16, value="获取失败")
                    else:
                        ws.cell(row=row_idx, column=16, value="格式错误")
                except Exception as e:
                    ws.cell(row=row_idx, column=16, value="处理出错")
            else:
                ws.cell(row=row_idx, column=16, value="空值")
            
            # 短暂等待
            time.sleep(0.3)
            
            processed_count += 1
        
        # 将处理后的工作簿保存到 BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回开头
        
        return success_count, output.getvalue()
        
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        traceback.print_exc()
        return 0, None

# 测试函数
def test_sequence_fetch():
    """测试序列获取功能"""
    test_input = "chr4B:425000640-425000640"
    print(f"测试获取序列: {test_input}")
    sequence = get_sequence_from_website_alternative(test_input)
    if sequence:
        print(f"成功获取序列，长度: {len(sequence)}")
        print(f"前50个字符: {sequence[:50]}")
    else:
        print("获取序列失败")
    return sequence

if __name__ == "__main__":
    # 运行测试
    test_sequence_fetch()
